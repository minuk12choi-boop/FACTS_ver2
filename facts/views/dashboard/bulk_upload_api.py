import csv
import io
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET, require_POST

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

from ...models import *
from ...permissions import _check_page_permission
from ..common import (
    _ensure_browser_close_session,
    _get_actor,
    _normalize_date_input,
    _normalize_upper,
    _plan_to_json,
    _tip_missing_to_json,
)
from .helpers import _is_invalid_single_cham_value

@require_POST
@login_required
def dashboard_bulk_upload_api(request):
    _ensure_browser_close_session(request)

    upload = request.FILES.get("file")
    snap_date_str = request.POST.get("snap_date")
    lineid = (request.POST.get("lineid") or "").strip()
    processid = (request.POST.get("processid") or request.POST.get("prp_processid") or "").strip()
    actor = _get_actor(request)
    permission_response = _check_page_permission(request, "dashboard", lineid=lineid, processid=processid, require_edit=True, popup=True)
    if permission_response is not None:
        return permission_response

    if not upload or not snap_date_str:
        return JsonResponse({"ok": False, "message": "파일과 기준일이 필요합니다."}, status=400)

    if not lineid:
        return JsonResponse({"ok": False, "message": "LINE 선택값이 필요합니다."}, status=400)

    snap_date = datetime.strptime(snap_date_str, "%Y-%m-%d").date()
    name = upload.name.lower()
    rows = []

    if name.endswith(".csv"):
        content = upload.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        for row_no, row in enumerate(reader, start=2):
            item = dict(row)
            item["__rownum__"] = row_no
            rows.append(item)
    elif name.endswith(".xlsx"):
        wb = load_workbook(upload, data_only=True)
        ws = wb["FACTS_UPLOAD_TEMPLATE"] if "FACTS_UPLOAD_TEMPLATE" in wb.sheetnames else wb.active
        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        for excel_row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            item = {}
            for idx, key in enumerate(header):
                item[key] = row[idx] if idx < len(row) else None
            item["__rownum__"] = excel_row_no
            rows.append(item)
    else:
        return JsonResponse({"ok": False, "message": "csv 또는 xlsx만 업로드 가능합니다."}, status=400)

    stage_map = {s.stage_code: s for s in FactsEvalStageMaster.objects.filter(is_active=True)}

    # 같은 파일 내 중복은 행번호가 작은 것만 반영
    seen_plan_keys = set()
    seen_tip_missing_keys = set()

    plan_applied = 0
    tip_applied = 0
    skipped_duplicate_in_file = 0
    skipped_invalid_cham = 0

    for r in rows:
        processid = _normalize_upper(r.get("PROCESSID"))
        stepseq = _normalize_upper(r.get("STEPSEQ"))

        plan_always = str(r.get("호환계획_상시/비상시") or "").strip()
        plan_major = str(r.get("호환계획_주요/비주요") or "").strip()
        plan_body = _normalize_upper(r.get("호환계획_호환EQPBODY명"))
        plan_cham = _normalize_upper(r.get("호환계획_호환EQPCHAM명"))
        plan_due = _normalize_date_input(r.get("호환계획_호환완료계획일"))
        eval_lot_id = _normalize_upper(r.get("호환계획_평가LotID"))
        stage_code = _normalize_upper(r.get("호환계획_평가단계"))
        memo = str(r.get("호환계획_비고") or "").strip()

        tip_missing_always = str(r.get("미등록TIP호환Path_상시/비상시") or "").strip()
        tip_missing_major = str(r.get("미등록TIP호환Path_주요/비주요") or "").strip()
        tip_missing_body = _normalize_upper(r.get("미등록TIP호환Path_호환EQPBODY명"))
        tip_missing_cham = _normalize_upper(r.get("미등록TIP호환Path_호환EQPCHAM명"))

        if not processid or not stepseq:
            continue

        plan_cham_invalid = _is_invalid_single_cham_value(plan_cham)
        tip_cham_invalid = _is_invalid_single_cham_value(tip_missing_cham)

        if plan_cham_invalid or tip_cham_invalid:
            skipped_invalid_cham += 1
            continue

        # 1) 호환계획 업로드
        if plan_body:
            plan_key = (lineid, processid, stepseq, plan_body, plan_cham)
            if plan_key in seen_plan_keys:
                skipped_duplicate_in_file += 1
            else:
                seen_plan_keys.add(plan_key)
                stage_obj = stage_map.get(stage_code) if stage_code else None

                existing_qs = FactsStepPlan.objects.filter(
                    lineid=lineid,
                    processid=processid,
                    stepseq=stepseq,
                    eqp_body_name=plan_body,
                    eqp_cham_name=plan_cham,
                    is_active=True,
                ).order_by("-updated_at", "-id")

                obj = existing_qs.first()
                if obj:
                    before_json = _plan_to_json(obj)

                    existing_qs.exclude(id=obj.id).update(
                        is_active=False,
                        updated_by=actor,
                    )

                    obj.always_emergency = plan_always
                    obj.major_minor = plan_major
                    obj.compatibility_due_date = plan_due
                    obj.eval_lot_id = eval_lot_id
                    obj.required_eval_stage = stage_obj
                    obj.memo = memo
                    obj.updated_by = actor
                    obj.save()

                    FactsEditHistory.objects.create(
                        action_type="bulk_upload",
                        snap_date=snap_date,
                        lineid=lineid,
                        processid=processid,
                        stepseq=stepseq,
                        recipeid=obj.recipeid or "",
                        changed_by=actor,
                        before_json=before_json,
                        after_json=_plan_to_json(obj),
                    )
                else:
                    obj = FactsStepPlan.objects.create(
                        lineid=lineid,
                        processid=processid,
                        stepseq=stepseq,
                        recipeid="",
                        always_emergency=plan_always,
                        major_minor=plan_major,
                        eqp_body_name=plan_body,
                        eqp_cham_name=plan_cham,
                        compatibility_due_date=plan_due,
                        eval_lot_id=eval_lot_id,
                        required_eval_stage=stage_obj,
                        memo=memo,
                        is_active=True,
                        created_by=actor,
                        updated_by=actor,
                    )

                    FactsEditHistory.objects.create(
                        action_type="bulk_upload",
                        snap_date=snap_date,
                        lineid=lineid,
                        processid=processid,
                        stepseq=stepseq,
                        recipeid="",
                        changed_by=actor,
                        before_json={},
                        after_json=_plan_to_json(obj),
                    )

                plan_applied += 1

        # 2) TIP미등록 호환Path 업로드
        if tip_missing_always and tip_missing_major and tip_missing_body:
            tip_key = (snap_date, lineid, processid, stepseq, tip_missing_body, tip_missing_cham)
            if tip_key in seen_tip_missing_keys:
                skipped_duplicate_in_file += 1
            else:
                seen_tip_missing_keys.add(tip_key)

                existing_tip_qs = FactsTipMissingCompatPath.objects.filter(
                    snap_date=snap_date,
                    lineid=lineid,
                    processid=processid,
                    stepseq=stepseq,
                    eqp_body_name=tip_missing_body,
                    eqp_cham_name=tip_missing_cham,
                    is_active=True,
                ).order_by("-updated_at", "-id")

                obj2 = existing_tip_qs.first()
                if obj2:
                    before_json = _tip_missing_to_json(obj2)

                    existing_tip_qs.exclude(id=obj2.id).update(
                        is_active=False,
                        updated_by=actor,
                    )

                    obj2.always_emergency = tip_missing_always
                    obj2.major_minor = tip_missing_major
                    obj2.updated_by = actor
                    obj2.save()

                    FactsEditHistory.objects.create(
                        action_type="bulk_upload",
                        snap_date=snap_date,
                        lineid=lineid,
                        processid=processid,
                        stepseq=stepseq,
                        recipeid=obj2.recipeid or "",
                        changed_by=actor,
                        before_json=before_json,
                        after_json=_tip_missing_to_json(obj2),
                    )
                else:
                    obj2 = FactsTipMissingCompatPath.objects.create(
                        snap_date=snap_date,
                        lineid=lineid,
                        processid=processid,
                        stepseq=stepseq,
                        recipeid="",
                        always_emergency=tip_missing_always,
                        major_minor=tip_missing_major,
                        eqp_body_name=tip_missing_body,
                        eqp_cham_name=tip_missing_cham,
                        is_active=True,
                        created_by=actor,
                        updated_by=actor,
                    )

                    FactsEditHistory.objects.create(
                        action_type="bulk_upload",
                        snap_date=snap_date,
                        lineid=lineid,
                        processid=processid,
                        stepseq=stepseq,
                        recipeid="",
                        changed_by=actor,
                        before_json={},
                        after_json=_tip_missing_to_json(obj2),
                    )

                tip_applied += 1

    return JsonResponse({
        "ok": True,
        "message": (
            f"업로드 완료. "
            f"호환계획 반영 {plan_applied}건, "
            f"TIP미등록 호환Path 반영 {tip_applied}건, "
            f"동일 파일 내 중복으로 스킵 {skipped_duplicate_in_file}건, "
            f"CHAM 입력 오류로 스킵 {skipped_invalid_cham}건"
        ),
    })

@require_GET
@login_required
def dashboard_upload_template(request):
    _ensure_browser_close_session(request)
    permission_response = _check_page_permission(request, "dashboard", ignore_blank_scope=True)
    if permission_response is not None:
        return permission_response

    wb = Workbook()
    ws = wb.active
    ws.title = "FACTS_UPLOAD_TEMPLATE"

    headers = [
        "PROCESSID",
        "STEPSEQ",
        "호환계획_상시/비상시",
        "호환계획_주요/비주요",
        "호환계획_호환EQPBODY명",
        "호환계획_호환EQPCHAM명",
        "호환계획_호환완료계획일",
        "호환계획_평가LotID",
        "호환계획_평가단계",
        "호환계획_비고",
        "미등록TIP호환Path_상시/비상시",
        "미등록TIP호환Path_주요/비주요",
        "미등록TIP호환Path_호환EQPBODY명",
        "미등록TIP호환Path_호환EQPCHAM명",
    ]
    ws.append(headers)

    ws.append([
        "P1SD",
        "SD00000000",
        "상시",
        "주요",
        "WSOD701",
        "F(*하나의 행엔 하나의 CHAM만 입력해주세요.)",
        "2026-04-13",
        "LOT123456",
        "",
        "계획 예시",
        "상시",
        "주요",
        "WSOD702",
        "1(*하나의 행엔 하나의 CHAM만 입력해주세요.)",
    ])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    header_map = {name: idx + 1 for idx, name in enumerate(headers)}

    widths = {
        "A": 16, "B": 16, "C": 25, "D": 25,
        "E": 25, "F": 30, "G": 25, "H": 18, "I": 16, "J": 24,
        "K": 30, "L": 30, "M": 34, "N": 34,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"

    dv_always = DataValidation(type="list", formula1='"상시,비상시"', allow_blank=True)
    dv_major = DataValidation(type="list", formula1='"주요,비주요"', allow_blank=True)

    stage_codes = list(
        FactsEvalStageMaster.objects.filter(is_active=True)
        .order_by("sort_order", "stage_code")
        .values_list("stage_code", flat=True)
    )
    stage_formula = '"' + ",".join(stage_codes) + '"' if stage_codes else '""'
    dv_stage = DataValidation(type="list", formula1=stage_formula, allow_blank=True)

    dv_date = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(2020,1,1)",
        formula2="DATE(2099,12,31)",
        allow_blank=True,
    )
    dv_date.error = "날짜는 엑셀 날짜 형식으로 입력하십시오. 예: 2026-04-13"
    dv_date.prompt = "호환완료계획일은 날짜 형식으로 입력"

    ws.add_data_validation(dv_always)
    ws.add_data_validation(dv_major)
    ws.add_data_validation(dv_stage)
    ws.add_data_validation(dv_date)

    dv_always.add("C2:C5000")
    dv_major.add("D2:D5000")
    dv_stage.add("I2:I5000")
    dv_date.add("G2:G5000")
    dv_always.add("K2:K5000")
    dv_major.add("L2:L5000")

    for row_idx in range(2, 5001):
        ws[f"G{row_idx}"].number_format = "yyyy-mm-dd"

    red_fill = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
        end_color="FFC7CE",
    )

    process_col = get_column_letter(header_map["PROCESSID"])
    step_col = get_column_letter(header_map["STEPSEQ"])
    plan_body_col = get_column_letter(header_map["호환계획_호환EQPBODY명"])
    plan_cham_col = get_column_letter(header_map["호환계획_호환EQPCHAM명"])
    tip_body_col = get_column_letter(header_map["미등록TIP호환Path_호환EQPBODY명"])
    tip_cham_col = get_column_letter(header_map["미등록TIP호환Path_호환EQPCHAM명"])

    # -------------------------------------------------
    # 숨김 helper 컬럼
    # AA: 호환계획 중복 key
    # AB: TIP미등록 중복 key
    # AC: 호환EQPCHAM명 유효성
    # AD: 미등록TIP호환Path_호환EQPCHAM명 유효성
    # -------------------------------------------------
    helper_cols = {
        "plan_dup_key": "AA",
        "tip_dup_key": "AB",
        "plan_cham_invalid": "AC",
        "tip_cham_invalid": "AD",
    }

    ws["AA1"] = "PLAN_DUP_KEY"
    ws["AB1"] = "TIP_DUP_KEY"
    ws["AC1"] = "PLAN_CHAM_INVALID"
    ws["AD1"] = "TIP_CHAM_INVALID"

    for row_idx in range(2, 5001):
        # CHAM이 비어 있으면 BODY까지만 key로 생성
        ws[f"AA{row_idx}"] = (
            f'=IF(TRIM(${plan_body_col}{row_idx})="", "", '
            f'UPPER(TRIM(${process_col}{row_idx}))&"|"&'
            f'UPPER(TRIM(${step_col}{row_idx}))&"|"&'
            f'UPPER(TRIM(${plan_body_col}{row_idx}))&"|"&'
            f'UPPER(TRIM(${plan_cham_col}{row_idx})))'
        )

        ws[f"AB{row_idx}"] = (
            f'=IF(TRIM(${tip_body_col}{row_idx})="", "", '
            f'UPPER(TRIM(${process_col}{row_idx}))&"|"&'
            f'UPPER(TRIM(${step_col}{row_idx}))&"|"&'
            f'UPPER(TRIM(${tip_body_col}{row_idx}))&"|"&'
            f'UPPER(TRIM(${tip_cham_col}{row_idx})))'
        )

        ws[f"AC{row_idx}"] = (
            f'=IF(TRIM(${plan_cham_col}{row_idx})="", FALSE, '
            f'OR(ISNUMBER(SEARCH(":",${plan_cham_col}{row_idx})),'
            f'ISNUMBER(SEARCH(";",${plan_cham_col}{row_idx})),'
            f'ISNUMBER(SEARCH(",",${plan_cham_col}{row_idx})),'
            f'LEN(TRIM(${plan_cham_col}{row_idx}))>1))'
        )

        ws[f"AD{row_idx}"] = (
            f'=IF(TRIM(${tip_cham_col}{row_idx})="", FALSE, '
            f'OR(ISNUMBER(SEARCH(":",${tip_cham_col}{row_idx})),'
            f'ISNUMBER(SEARCH(";",${tip_cham_col}{row_idx})),'
            f'ISNUMBER(SEARCH(",",${tip_cham_col}{row_idx})),'
            f'LEN(TRIM(${tip_cham_col}{row_idx}))>1))'
        )

    # helper 컬럼 숨김
    for col in ["AA", "AB", "AC", "AD"]:
        ws.column_dimensions[col].hidden = True

    # -------------------------------------------------
    # 조건부서식
    # -------------------------------------------------

    # 1) 호환계획 중복 경고 (BODY만 있어도 key 생성되므로 잡힘)
    ws.conditional_formatting.add(
        f"{plan_body_col}2:{plan_cham_col}5000",
        FormulaRule(
            formula=['=AND($AA2<>"",COUNTIF($AA:$AA,$AA2)>1)'],
            fill=red_fill,
        ),
    )

    # 2) TIP미등록 중복 경고
    ws.conditional_formatting.add(
        f"{tip_body_col}2:{tip_cham_col}5000",
        FormulaRule(
            formula=['=AND($AB2<>"",COUNTIF($AB:$AB,$AB2)>1)'],
            fill=red_fill,
        ),
    )

    # 3) STEPSEQ 소문자 경고
    ws.conditional_formatting.add(
        f"{step_col}2:{step_col}5000",
        FormulaRule(
            formula=[f'=AND(${step_col}2<>"",EXACT(${step_col}2,UPPER(${step_col}2))=FALSE)'],
            fill=red_fill,
        ),
    )

    # 4) 호환EQPCHAM명 다중입력/1글자초과 경고
    ws.conditional_formatting.add(
        f"{plan_cham_col}2:{plan_cham_col}5000",
        FormulaRule(
            formula=['=$AC2=TRUE'],
            fill=red_fill,
        ),
    )

    # 5) 미등록TIP호환Path_호환EQPCHAM명 다중입력/1글자초과 경고
    ws.conditional_formatting.add(
        f"{tip_cham_col}2:{tip_cham_col}5000",
        FormulaRule(
            formula=['=$AD2=TRUE'],
            fill=red_fill,
        ),
    )

    guide_ws = wb.create_sheet("작성 참고")
    guide_ws["A1"] = "FACTS 엑셀 업로드 작성 참고"
    guide_ws["A1"].font = Font(bold=True, size=13)

    guide_rows = [
        "1. 업로드는 FACTS_UPLOAD_TEMPLATE 시트만 읽습니다.",
        "2. 동일 파일 내에서 같은 설비가 중복되면 행번호가 더 작은 행만 반영됩니다.",
        "3. 같은 설비가 기존 DB에 이미 있으면 새 업로드 값으로 덮어씌웁니다.",
        "4. 중복 경고 기준",
        "   - 호환계획: PROCESSID + STEPSEQ + 호환계획_호환EQPBODY명 + 호환계획_호환EQPCHAM명",
        "   - TIP미등록: PROCESSID + STEPSEQ + 미등록TIP호환Path_호환EQPBODY명 + 미등록TIP호환Path_호환EQPCHAM명",
        "5. CHAM명은 비어 있을 수 있으며, 비어 있어도 BODY 기준으로 중복 경고가 동작합니다.",
        "6. PROCESSID, STEPSEQ, 호환계획_호환EQPBODY명, 호환계획_호환EQPCHAM명, 호환계획_평가LotID, 미등록TIP호환Path_호환EQPBODY명, 미등록TIP호환Path_호환EQPCHAM명은 업로드 시 서버에서 영문자를 무조건 대문자로 변환합니다.",
        "7. 하나의 행에는 하나의 CHAM단위까지만 기입 가능합니다. 하나 이상의 CHAM 기입(예시: 3:4, 3;4, 3,4 등) 된 행은 업데이트가 안됩니다.",
        "8. 호환계획_호환EQPCHAM명, 미등록TIP호환Path_호환EQPCHAM명 컬럼에 다중 CHAM 형식 또는 1글자 초과가 입력되면 빨간색 warning으로 표시됩니다.",
        "9. STEPSEQ에 소문자가 입력되면 빨간색 warning으로 표시됩니다.",
        "10. 상시/비상시는 해당 호환 계획 혹은 미등록TIP호환Path가 상시 호환 가능한 것인지 비상시에만 호환 가능한 것인지 선택하는 것입니다.",
        "11. 주요/비주요는 해당 호환 계획 혹은 미등록TIP호환Path가 주요랏도 갈 수 있을 때엔 주요를 선택하시고, 비주요랏만 갈 수 있는 경우엔 비주요를 선택바랍니다.",
    ]

    for idx, text in enumerate(guide_rows, start=3):
        guide_ws[f"A{idx}"] = text

    guide_ws.column_dimensions["A"].width = 140

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="facts_upload_template.xlsx"'
    return response

